from pathlib import Path
from openpyxl import load_workbook

class ExcelFetcher:

    def __init__(self):
        self.excel_file = self.resolve_excel_path()
        self.rows = self.read_col_c_d(self.excel_file)
    # this function gets the path of Excel file located in project
    @staticmethod
    def resolve_excel_path():
        # Looking for a file locally
        # It has to be seated in 'excel' folder
        project_root = Path(__file__).resolve().parents[1]

        # sets a path where excel should be located
        excel_dir = project_root / "excel"

        # gets a full path of the first xlsx file
        files = list(excel_dir.glob("*.xlsx"))
        if not files:
            raise FileNotFoundError(f"No excel files in {excel_dir}")
        first_file = files[0]
        return first_file

    # this function reads the data in rows
    # REQUIREMENTS
    # column 'C' from second row: Job Title
    # column 'D' from second row: Name of employee
    @staticmethod
    def read_col_c_d(file_path):
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if "Reference User" not in wb.sheetnames:
            raise KeyError("Sheet 'Reference User' not found")
        ws = wb["Reference User"]

        # read from second row in column C and D
        result = []
        for c_val, d_val in ws.iter_rows(min_row=2, min_col=3, max_col=4, values_only=True):
            # c_val is column C, d_val is column D
            if c_val is None and d_val is None:
                continue
            result.append((c_val, d_val))
        wb.close()
        return result

    excel_file = resolve_excel_path()
    # rows is a tuple: ('Job Title', 'Firstname Lastname')
    rows = read_col_c_d(excel_file)
